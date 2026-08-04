"""
Adaptador de Parser para a Prefeitura de São Paulo.
Extrai notas fiscais do Registro de Notas Fiscais de Serviços Prestados (Mod. 51) de São Paulo.
"""

import pdfplumber
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser

class SaoPauloParser(BaseCityParser):
    def __init__(self):
        super().__init__("São Paulo")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        records = []
        
        if isinstance(file_source, bytes):
            if file_source.startswith(b'%PDF'):
                return self._parse_pdf_bytes(file_source)
            elif file_source.startswith(b'PK'):
                return self._parse_xlsx_bytes(file_source)
            else:
                return self._parse_csv_bytes(file_source)
        elif isinstance(file_source, str):
            if file_source.lower().endswith('.pdf'):
                with open(file_source, 'rb') as f:
                    return self._parse_pdf_bytes(f.read())
            elif file_source.lower().endswith('.xlsx'):
                with open(file_source, 'rb') as f:
                    return self._parse_xlsx_bytes(f.read())
            else:
                with open(file_source, 'rb') as f:
                    return self._parse_csv_bytes(f.read())
        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        current_city = "São Paulo"

        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text:
                        continue

                    for line in text.split('\n'):
                        if "SAO PAULO/SP" in line or "SÃO PAULO/SP" in line:
                            current_city = "São Paulo"
                        
                        if '|' not in line:
                            continue

                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0:
                                        continue

                                    val_iss = 0.0
                                    if len(parts) >= 6:
                                        iss_raw = parts[5].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    records.append({
                                        "id": f"SP-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": str(int(numero)) if numero.isdigit() else numero,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "São Paulo"
                                    })
                                    idx += 1
                                except ValueError:
                                    pass
        except Exception:
            pass

        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = csv_bytes.decode('latin1', errors='replace')
            lines = content.splitlines()
            reader = csv.reader(lines, delimiter=';')
            headers = next(reader, None)
            if not headers: return []

            idx_situacao = None
            idx_valor = None
            idx_valor_iss = None
            idx_numero = None

            for idx, h in enumerate(headers):
                h_norm = h.lower()
                if 'situa' in h_norm and 'nota' in h_norm:
                    idx_situacao = idx
                elif 'valor dos servi' in h_norm or 'base' in h_norm:
                    if idx_valor is None: idx_valor = idx
                elif 'iss devido' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'nfs' in h_norm or 'numero' in h_norm or 'nota' in h_norm:
                    if idx_numero is None: idx_numero = idx

            if idx_numero is None: idx_numero = 1
            if idx_valor is None: idx_valor = 26
            if idx_valor_iss is None: idx_valor_iss = 30
            if idx_situacao is None: idx_situacao = 22

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= (idx_valor or 0): continue

                # Ignora linha de Total/Rodapé consolidado
                first_cell = row[0].strip().lower() if len(row) > 0 else ""
                if first_cell.startswith('total') or first_cell.startswith('soma'):
                    continue

                # Filtro estrito de Situação 'T' (Tributada) - Ignora 'C' e vazias
                if idx_situacao is not None:
                    sit = row[idx_situacao].strip().upper() if idx_situacao < len(row) else ''
                    if sit != 'T':
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                if not raw_val: continue

                try:
                    val = float(raw_val.replace('.', '').replace(',', '.'))
                    if val <= 0: continue
                    
                    val_iss = 0.0
                    if idx_valor_iss is not None and idx_valor_iss < len(row):
                        raw_iss = row[idx_valor_iss].strip()
                        if raw_iss:
                            try:
                                iss_str = raw_iss.replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
                                if iss_str:
                                    val_iss = float(iss_str.replace('.', '').replace(',', '.'))
                            except ValueError: pass
                    
                    num_cell = row[idx_numero].strip() if idx_numero and idx_numero < len(row) else f"SP-{idx_row+1}"
                    num = str(int(num_cell)) if num_cell.isdigit() else num_cell

                    records.append({
                        "id": f"SP-{len(records)+1}",
                        "linha": idx_row + 1,
                        "numero": num,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": raw_val,
                        "cidade": "São Paulo"
                    })
                except ValueError:
                    pass
        except Exception:
            pass
        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        import openpyxl
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_situacao = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if ('nfs-e' in h_norm or 'nfse' in h_norm or 'numero' in h_norm) and idx_numero is None:
                    idx_numero = idx
                elif 'valor dos servicos' in h_norm or 'valor servicos' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'iss devido' in h_norm or 'valor do iss' in h_norm or h_norm == 'iss':
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'situacao' in h_norm or 'status' in h_norm:
                    idx_situacao = idx

            if idx_numero is None: idx_numero = 1
            if idx_valor is None: idx_valor = 26
            if idx_valor_iss is None: idx_valor_iss = 30
            if idx_situacao is None: idx_situacao = 22

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                sit_cell = sheet.cell(row=row_idx, column=idx_situacao+1).value if idx_situacao is not None else 'T'

                sit_str = str(sit_cell).strip().upper() if sit_cell is not None else ''
                if idx_situacao is not None and sit_str != 'T':
                    continue

                if val_cell is None: continue
                try:
                    val_str = str(val_cell).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
                    if not val_str: continue
                    val = float(val_str.replace('.', '').replace(',', '.'))
                    if val <= 0: continue

                    val_iss = 0.0
                    if idx_valor_iss is not None:
                        iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                        if iss_cell is not None:
                            try:
                                iss_str = str(iss_cell).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
                                if iss_str:
                                    val_iss = float(iss_str.replace('.', '').replace(',', '.'))
                            except ValueError: pass

                    nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell or '').strip()

                    records.append({
                        "id": f"SP-{len(records)+1}",
                        "linha": row_idx,
                        "numero": nf,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": str(val_cell),
                        "cidade": "São Paulo"
                    })
                except ValueError:
                    pass
        except Exception:
            pass

        return records
