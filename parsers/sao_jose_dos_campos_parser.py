"""
Adaptador de Parser para a Prefeitura de São José dos Campos.
Extrai notas fiscais do Registro de Notas Fiscais de Serviços Prestados de São José dos Campos (PDF ou CSV).
"""

import pdfplumber
import io
import os
import csv
import openpyxl
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser, safe_read_bytes

def parse_val(val_raw) -> float:
    if isinstance(val_raw, (int, float)):
        return float(val_raw)
    val_str = str(val_raw).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

class SaoJoseDosCamposParser(BaseCityParser):
    def __init__(self):
        super().__init__("São José dos Campos")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        records = []
        
        if isinstance(file_source, bytes):
            if file_source.startswith(b'%PDF'):
                return self._parse_pdf_bytes(file_source)
            elif file_source.startswith(b'PK'):
                return self._parse_xlsx_bytes(file_source)
            else:
                return self._parse_csv_bytes(file_source)
        elif isinstance(file_source, str):
            if file_source.lower().endswith('.pdf'):
                with open(file_source, 'rb') as f:
                    return self._parse_pdf_bytes(f.read())
            elif file_source.lower().endswith('.xlsx'):
                with open(file_source, 'rb') as f:
                    return self._parse_xlsx_bytes(f.read())
            else:
                with open(file_source, 'rb') as f:
                    return self._parse_csv_bytes(f.read())
        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        current_city = "São José dos Campos"

        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text:
                        continue

                    for line in text.split('\n'):
                        if "SAO JOSE DOS CAMPOS" in line or "SÃO JOSÉ DOS CAMPOS" in line:
                            current_city = "São José dos Campos"
                        
                        if '|' not in line:
                            continue

                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0:
                                        continue

                                    val_iss = 0.0
                                    if len(parts) >= 6:
                                        iss_raw = parts[5].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    records.append({
                                        "id": f"SJC-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": numero,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "São José dos Campos"
                                    })
                                    idx += 1
                                except ValueError:
                                    pass
        except Exception:
            pass

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_cancelamento = None
            idx_tomador = None
            idx_retido = None
            idx_situacao = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'nfs' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'valor servicos' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif h_norm == 'iss' or 'issqn' in h_norm or 'valor iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'iss retido' in h_norm:
                    idx_retido = idx
                elif 'cancelamento' in h_norm:
                    idx_cancelamento = idx
                elif 'situa' in h_norm:
                    idx_situacao = idx
                elif 'tomador - nome' in h_norm or ('razao' in h_norm and idx_tomador is None):
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8
            if idx_cancelamento is None: idx_cancelamento = 11

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                canc_cell = sheet.cell(row=row_idx, column=idx_cancelamento+1).value if idx_cancelamento is not None else None
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value if idx_tomador is not None else None

                # NFS-e Nacional cancelamento
                if canc_cell and str(canc_cell).strip():
                    continue

                # Legacy situacao check
                if idx_situacao is not None:
                    sit = sheet.cell(row=row_idx, column=idx_situacao+1).value
                    if sit and str(sit).strip().upper() not in ['ATIVA', 'ATIVO', 'T', 'TRIBUTADA']:
                        continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                val_iss = 0.0
                if idx_valor_iss is not None:
                    iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                    if iss_cell is not None:
                        val_iss = parse_val(iss_cell)

                retido_val = ""
                if idx_retido is not None:
                    r_cell = sheet.cell(row=row_idx, column=idx_retido+1).value
                    if r_cell is not None:
                        retido_val = str(r_cell).strip().lower()
                
                iss_ret_flag = "S" if retido_val == "sim" else "N"

                records.append({
                    "id": f"SJC-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "iss_retido": iss_ret_flag,
                    "cidade": "São José dos Campos"
                })
        except Exception:
            pass

        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content_str = ""
            for enc in ['latin1', 'utf-8', 'cp1252']:
                try:
                    content_str = csv_bytes.decode(enc)
                    break
                except Exception:
                    pass

            lines = content_str.splitlines()
            if not lines: return []

            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            rows_list = list(reader)

            headers = None
            header_row_idx = -1

            for idx, row in enumerate(rows_list):
                row_str = " ".join(row).lower()
                if 'valor' in row_str or 'nfs-e nacional' in row_str or 'situa' in row_str:
                    headers = row
                    header_row_idx = idx
                    break

            if not headers:
                headers = rows_list[0] if rows_list else []
                header_row_idx = 0

            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_tomador = None
            idx_retido = None
            idx_situacao = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if 'nfs-e nacional' in h_norm:
                    idx_numero = idx
                elif ('nota fiscal' in h_norm or 'nfs' in h_norm or 'nr. nf' in h_norm or 'numero' in h_norm) and idx_numero is None:
                    idx_numero = idx
                elif 'valor' in h_norm and ('servi' in h_norm or idx_valor is None):
                    idx_valor = idx
                elif 'iss retido' in h_norm:
                    idx_retido = idx
                elif 'issqn' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'raz' in h_norm and 'tomador' in h_norm:
                    idx_tomador = idx
                elif 'situa' in h_norm:
                    idx_situacao = idx

            if idx_numero is None: idx_numero = 1
            if idx_valor is None: idx_valor = 9

            for idx_row in range(header_row_idx + 1, len(rows_list)):
                row = rows_list[idx_row]
                if not row or len(row) <= idx_valor:
                    continue
                
                if row[0].strip().lower() == 'total':
                    continue

                if idx_situacao is not None and idx_situacao < len(row):
                    sit = row[idx_situacao].strip().upper()
                    if sit and sit not in ['ATIVA', 'ATIVO', 'T', 'TRIBUTADA']:
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                if not raw_val: continue
                val = parse_val(raw_val)
                if val <= 0: continue

                val_iss = 0.0
                if idx_valor_iss is not None and idx_valor_iss < len(row):
                    raw_iss = row[idx_valor_iss].strip()
                    if raw_iss:
                        val_iss = parse_val(raw_iss)

                num_cell = row[idx_numero].strip() if idx_numero < len(row) else ""
                if not num_cell and len(row) > 1:
                    num_cell = row[1].strip()
                if not num_cell:
                    num_cell = f"SJC-{idx_row+1}"
                
                nf = str(int(num_cell)) if num_cell.isdigit() else num_cell
                tomador = row[idx_tomador].strip() if idx_tomador is not None and idx_tomador < len(row) else ""
                
                retido_val = row[idx_retido].strip() if idx_retido is not None and idx_retido < len(row) else ""
                iss_ret_flag = "S" if retido_val == "1" or retido_val.lower() == "sim" else "N"

                records.append({
                    "id": f"SJC-{len(records)+1}",
                    "linha": idx_row + 1,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": raw_val,
                    "tomador": tomador,
                    "iss_retido": iss_ret_flag,
                    "cidade": "São José dos Campos"
                })
        except Exception:
            pass

        return records
