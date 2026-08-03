"""
Adaptador de Parser para a Prefeitura de Santos.
Extrai notas fiscais do Registro de Notas Fiscais de Serviços Prestados de Santos (PDF ou XLSX/CSV).
"""

import pdfplumber
import openpyxl
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser

class SantosParser(BaseCityParser):
    def __init__(self):
        super().__init__("Santos")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        records = []
        
        if isinstance(file_source, bytes):
            if file_source.startswith(b'%PDF'):
                return self._parse_pdf_bytes(file_source)
            elif file_source.startswith(b'PK'):
                return self._parse_xlsx_bytes(file_source)
            else:
                return self._parse_csv_bytes(file_source)
        elif isinstance(file_source, str):
            if file_source.lower().endswith('.pdf'):
                with open(file_source, 'rb') as f:
                    return self._parse_pdf_bytes(f.read())
            elif file_source.lower().endswith('.xlsx') or file_source.lower().endswith('.xls'):
                with open(file_source, 'rb') as f:
                    return self._parse_xlsx_bytes(f.read())
            else:
                with open(file_source, 'rb') as f:
                    return self._parse_csv_bytes(f.read())
        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text:
                        continue

                    for line in text.split('\n'):
                        if '|' not in line:
                            continue

                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0:
                                        continue

                                    val_iss = 0.0
                                    if len(parts) >= 7:
                                        iss_raw = parts[6].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    num_clean = str(int(numero))

                                    records.append({
                                        "id": f"SAN-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "Santos"
                                    })
                                    idx += 1
                                except ValueError:
                                    pass
        except Exception:
            pass

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_escrituracao = None
            idx_valor = None
            idx_valor_iss = None
            idx_nfs = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'escrituracao' in h_norm:
                    idx_escrituracao = idx
                elif 'valor do servico' in h_norm or ('base' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'issqn' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'nfs' in h_norm and idx_nfs is None:
                    idx_nfs = idx
                elif 'nome/razao social' in h_norm or ('razao' in h_norm and idx_tomador is None):
                    if idx > 13:
                        idx_tomador = idx

            if idx_escrituracao is None: idx_escrituracao = 41
            if idx_valor is None: idx_valor = 23
            if idx_nfs is None: idx_nfs = 3
            if idx_tomador is None: idx_tomador = 16

            for row_idx in range(2, sheet.max_row+1):
                esc = sheet.cell(row=row_idx, column=idx_escrituracao+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                nfs_cell = sheet.cell(row=row_idx, column=idx_nfs+1).value
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value

                if esc and str(esc).strip().upper() not in ['ATIVA', 'ATIVO', 'T']:
                    continue

                if val_cell is None: continue
                try:
                    val = float(val_cell)
                    if val <= 0: continue

                    val_iss = 0.0
                    if idx_valor_iss is not None:
                        iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                        if iss_cell is not None:
                            try:
                                iss_str = str(iss_cell).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
                                if iss_str:
                                    val_iss = float(iss_str.replace('.', '').replace(',', '.'))
                            except ValueError: pass

                    nfs = str(int(nfs_cell)) if isinstance(nfs_cell, (int, float)) else str(nfs_cell)

                    records.append({
                        "id": f"SAN-{len(records)+1}",
                        "linha": row_idx,
                        "numero": nfs,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": str(val_cell),
                        "tomador": str(tomador_cell or ''),
                        "cidade": "Santos"
                    })
                except ValueError:
                    pass
        except Exception:
            pass

        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = csv_bytes.decode('utf-8', errors='replace')
            lines = content.splitlines()
            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers: return []

            idx_escrituracao = None
            idx_valor = None
            idx_valor_iss = None
            idx_nfs = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'escrituracao' in h_norm:
                    idx_escrituracao = idx
                elif 'valor do servico' in h_norm or ('base' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'issqn' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'nfs' in h_norm and idx_nfs is None:
                    idx_nfs = idx

            if idx_escrituracao is None: idx_escrituracao = 41
            if idx_valor is None: idx_valor = 23
            if idx_nfs is None: idx_nfs = 3

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= idx_valor: continue

                if idx_escrituracao < len(row):
                    esc = row[idx_escrituracao].strip().upper()
                    if esc and esc not in ['ATIVA', 'ATIVO', 'T']:
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                if not raw_val: continue

                try:
                    val = float(raw_val.replace('.', '').replace(',', '.'))
                    if val <= 0: continue
                    
                    val_iss = 0.0
                    if idx_valor_iss is not None and idx_valor_iss < len(row):
                        raw_iss = row[idx_valor_iss].strip()
                        if raw_iss:
                            try:
                                iss_str = raw_iss.replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
                                if iss_str:
                                    val_iss = float(iss_str.replace('.', '').replace(',', '.'))
                            except ValueError: pass
                    
                    num = row[idx_nfs].strip() if idx_nfs < len(row) else f"SAN-{idx_row+1}"

                    records.append({
                        "id": f"SAN-{len(records)+1}",
                        "linha": idx_row + 1,
                        "numero": num,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": raw_val,
                        "cidade": "Santos"
                    })
                except ValueError:
                    pass
        except Exception:
            pass
        return records
