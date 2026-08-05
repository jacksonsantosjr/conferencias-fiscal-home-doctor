"""
Adaptador de Parser para a Prefeitura de Volta Redonda (RJ).
Extrai notas fiscais do relatório oficial da Prefeitura de Volta Redonda (XLS, XLSX, CSV ou PDF),
com suporte para leitura de planilhas .xls legadas (BIFF) via xlrd, filtragem de notas canceladas e extração da coluna 'Valor Servicos'.
"""

import pdfplumber
import openpyxl
import xlrd
import io
import os
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser, safe_read_bytes

def parse_val(val_raw) -> float:
    if isinstance(val_raw, (int, float)):
        return float(val_raw)
    val_str = str(val_raw).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

class VoltaRedondaParser(BaseCityParser):
    def __init__(self):
        super().__init__("Volta Redonda")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)
        elif data_bytes.startswith(b'\xd0\xcf\x11\xe0'):
            return self._parse_xls_bytes(data_bytes)
        else:
            return self._parse_csv_bytes(data_bytes)

    def _parse_xls_bytes(self, xls_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = xlrd.open_workbook(file_contents=xls_bytes)
            sheet = wb.sheet_by_index(0)

            header_row = 1
            for r in range(min(5, sheet.nrows)):
                row_str = [str(sheet.cell_value(r, c)).lower() for c in range(sheet.ncols)]
                if any('numero' in s or 'número' in s for s in row_str):
                    header_row = r
                    break

            headers = [str(sheet.cell_value(header_row, c)).strip() for c in range(sheet.ncols)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_status = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'valor servicos' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'valor iss' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'status' in h_norm or 'situacao' in h_norm:
                    idx_status = idx
                elif 'tomador' in h_norm:
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 1
            if idx_valor is None: idx_valor = 15
            if idx_status is None: idx_status = 11
            if idx_tomador is None: idx_tomador = 8

            for row_idx in range(header_row + 1, sheet.nrows):
                status_val = str(sheet.cell_value(row_idx, idx_status)).strip().upper() if idx_status < sheet.ncols else ''
                if status_val and status_val in ['CANCELADA', 'CANCELADO']:
                    continue

                num_cell = sheet.cell_value(row_idx, idx_numero)
                val_cell = sheet.cell_value(row_idx, idx_valor)
                tomador_cell = sheet.cell_value(row_idx, idx_tomador) if idx_tomador < sheet.ncols else ''

                if val_cell is None or val_cell == '': continue

                val_iss = 0.0
                if idx_valor_iss is not None and idx_valor_iss < sheet.ncols:
                    iss_cell = sheet.cell_value(row_idx, idx_valor_iss)
                    if iss_cell is not None and iss_cell != '':
                        try:
                            val_iss = float(str(iss_cell).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) if isinstance(iss_cell, str) else float(iss_cell)
                        except ValueError:
                            pass

                try:
                    val = float(str(val_cell).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) if isinstance(val_cell, str) else float(val_cell)
                    if val <= 0: continue

                    nf = str(int(num_cell)) if isinstance(num_cell, float) and num_cell.is_integer() else str(num_cell).strip()
                    if nf.endswith('.0'): nf = nf[:-2]

                    records.append({
                        "id": f"VR-{len(records)+1}",
                        "linha": row_idx + 1,
                        "numero": nf,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": str(val_cell),
                        "tomador": str(tomador_cell or ''),
                        "cidade": "Volta Redonda"
                    })
                except ValueError:
                    pass
        except Exception:
            pass

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_cancelamento = None
            idx_tomador = None
            idx_retido = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'nfs' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'valor servicos' in h_norm or 'vl. nf' in h_norm or 'valor nf' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif h_norm == 'iss' or 'issqn' in h_norm or 'vl. iss' in h_norm or 'valor iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'iss retido' in h_norm:
                    idx_retido = idx
                elif 'cancelamento' in h_norm or 'status' in h_norm or 'situacao' in h_norm:
                    idx_cancelamento = idx
                elif 'tomador' in h_norm:
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8
            if idx_cancelamento is None: idx_cancelamento = 11

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                canc_cell = sheet.cell(row=row_idx, column=idx_cancelamento+1).value if idx_cancelamento is not None else None
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value if idx_tomador is not None else None

                # NFS-e Nacional cancelamento
                if canc_cell and (str(canc_cell).strip().upper() in ['CANCELADA', 'C', 'CANCELADO'] or (idx_cancelamento == 11 and str(canc_cell).strip() != "")):
                    continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                val_iss = 0.0
                if idx_valor_iss is not None:
                    iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                    if iss_cell is not None:
                        val_iss = parse_val(iss_cell)

                retido_val = ""
                if idx_retido is not None:
                    r_cell = sheet.cell(row=row_idx, column=idx_retido+1).value
                    if r_cell is not None:
                        retido_val = str(r_cell).strip().lower()
                
                iss_ret_flag = "S" if retido_val == "sim" else "N"

                records.append({
                    "id": f"VR-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "iss_retido": iss_ret_flag,
                    "cidade": "Volta Redonda"
                })
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        if '|' not in line: continue
                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue
                                    
                                    val_iss = 0.0
                                    if len(parts) >= 6:
                                        iss_raw = parts[5].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"VR-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "Volta Redonda"
                                    })
                                    idx += 1
                                except ValueError: pass
        except Exception: pass
        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content_str = ""
            for enc in ['latin1', 'utf-8', 'cp1252']:
                try:
                    content_str = csv_bytes.decode(enc)
                    break
                except Exception:
                    pass

            lines = content_str.splitlines()
            if not lines: return []

            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            rows_list = list(reader)

            headers = None
            header_row_idx = -1

            for idx, row in enumerate(rows_list):
                row_str = " ".join(row).lower()
                if 'valor' in row_str or 'nfs-e nacional' in row_str:
                    headers = row
                    header_row_idx = idx
                    break

            if not headers:
                headers = rows_list[0] if rows_list else []
                header_row_idx = 0

            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_tomador = None
            idx_retido = None
            idx_cancelamento = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if 'nfs-e nacional' in h_norm:
                    idx_numero = idx
                elif ('nota fiscal' in h_norm or 'nfs' in h_norm or 'nr. nf' in h_norm or 'numero' in h_norm) and idx_numero is None:
                    idx_numero = idx
                elif 'vl. nf' in h_norm or 'valor nf' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'iss retido' in h_norm:
                    idx_retido = idx
                elif 'vl. iss' in h_norm or 'issqn' in h_norm or 'iss' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'raz' in h_norm and 'tomador' in h_norm:
                    idx_tomador = idx
                elif 'cancelamento' in h_norm or 'status' in h_norm or 'situacao' in h_norm:
                    idx_cancelamento = idx

            if idx_numero is None: idx_numero = 1
            if idx_valor is None: idx_valor = 9

            for idx_row in range(header_row_idx + 1, len(rows_list)):
                row = rows_list[idx_row]
                if not row or len(row) <= idx_valor:
                    continue
                
                if row[0].strip().lower() == 'total':
                    continue

                if idx_cancelamento is not None and idx_cancelamento < len(row):
                    sit = row[idx_cancelamento].strip().upper()
                    if sit in ['CANCELADA', 'C', 'CANCELADO']:
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                if not raw_val: continue
                val = parse_val(raw_val)
                if val <= 0: continue

                val_iss = 0.0
                if idx_valor_iss is not None and idx_valor_iss < len(row):
                    raw_iss = row[idx_valor_iss].strip()
                    if raw_iss:
                        val_iss = parse_val(raw_iss)

                num_cell = row[idx_numero].strip() if idx_numero < len(row) else ""
                if not num_cell and len(row) > 1:
                    num_cell = row[1].strip()
                if not num_cell:
                    num_cell = f"VR-{idx_row+1}"
                
                nf = str(int(num_cell)) if num_cell.isdigit() else num_cell
                tomador = row[idx_tomador].strip() if idx_tomador is not None and idx_tomador < len(row) else ""
                
                retido_val = row[idx_retido].strip() if idx_retido is not None and idx_retido < len(row) else ""
                iss_ret_flag = "S" if retido_val == "1" or retido_val.lower() == "sim" else "N"

                records.append({
                    "id": f"VR-{len(records)+1}",
                    "linha": idx_row + 1,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": raw_val,
                    "tomador": tomador,
                    "iss_retido": iss_ret_flag,
                    "cidade": "Volta Redonda"
                })
        except Exception:
            pass
        return records
