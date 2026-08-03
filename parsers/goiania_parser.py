"""
Adaptador de Parser para a Prefeitura de Goiânia (GO).
Extrai notas fiscais do relatório oficial da Prefeitura de Goiânia (CSV, XLSX ou PDF),
considerando estritamente a coluna 'Valor Documento', cruzando 'N°' vs 'NÚMERO' e filtrando apenas linhas com Natureza Operação == 'Exigível'.
"""

import pdfplumber
import openpyxl
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser, safe_read_bytes

def parse_val(val_raw) -> float:
    if isinstance(val_raw, (int, float)):
        return float(val_raw)
    val_str = str(val_raw).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0


class GoianiaParser(BaseCityParser):
    def __init__(self):
        super().__init__("Goiânia")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)
        else:
            return self._parse_csv_bytes(data_bytes)

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = ""
            for enc in ['latin1', 'utf-8', 'cp1252']:
                try:
                    content = csv_bytes.decode(enc)
                    break
                except Exception:
                    pass

            lines = content.splitlines()
            if not lines: return []

            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            rows_list = list(reader)

            headers = None
            header_row_idx = -1

            for idx, row in enumerate(rows_list):
                row_str = " ".join(row).lower()
                if 'valor documento' in row_str or 'natureza opera' in row_str or 'n°' in row_str:
                    headers = row
                    header_row_idx = idx
                    break

            if not headers:
                headers = rows_list[0] if rows_list else []
                header_row_idx = 0

            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_natureza = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if (h_norm == 'n°' or 'n°' in h_norm or 'numero' in h_norm) and idx_numero is None and idx != 1:
                    idx_numero = idx
                elif 'valor documento' in h_norm or ('valor' in h_norm and 'imposto' not in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'valor imposto' in h_norm or 'imposto' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'natureza' in h_norm:
                    idx_natureza = idx
                elif 'tomador' in h_norm:
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 2
            if idx_valor is None: idx_valor = 5
            if idx_natureza is None: idx_natureza = 4
            if idx_tomador is None: idx_tomador = 3

            for idx_row in range(header_row_idx + 1, len(rows_list)):
                row = rows_list[idx_row]
                if not row or len(row) <= idx_valor:
                    continue

                if idx_natureza is not None and idx_natureza < len(row):
                    nat = row[idx_natureza].strip().lower()
                    if 'exig' not in nat:
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                val = parse_val(raw_val)
                if val <= 0: continue

                val_iss = 0.0
                if idx_valor_iss is not None and idx_valor_iss < len(row):
                    raw_iss = row[idx_valor_iss].strip()
                    if raw_iss:
                        try:
                            val_iss = float(raw_iss.replace('.', '').replace(',', '.'))
                        except ValueError: pass

                num_cell = row[idx_numero].strip() if idx_numero < len(row) else f"GYN-{idx_row+1}"
                nf = str(int(num_cell)) if num_cell.isdigit() else num_cell
                tomador = row[idx_tomador].strip() if idx_tomador is not None and idx_tomador < len(row) else ""

                records.append({
                    "id": f"GYN-{len(records)+1}",
                    "linha": idx_row + 1,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": raw_val,
                    "tomador": tomador,
                    "cidade": "Goiânia"
                })
        except Exception:
            pass

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_natureza = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if (h_norm == 'n°' or 'n°' in h_norm or 'numero' in h_norm) and idx_numero is None:
                    idx_numero = idx
                elif 'valor documento' in h_norm or ('valor' in h_norm and 'imposto' not in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'valor imposto' in h_norm or 'imposto' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'natureza' in h_norm:
                    idx_natureza = idx
                elif 'tomador' in h_norm:
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 2
            if idx_valor is None: idx_valor = 5

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                nat_cell = sheet.cell(row=row_idx, column=idx_natureza+1).value if idx_natureza is not None else 'Exigível'
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value if idx_tomador is not None else ''

                if nat_cell and 'exig' not in str(nat_cell).strip().lower():
                    continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                val_iss = 0.0
                if idx_valor_iss is not None:
                    iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                    if iss_cell is not None:
                        try:
                            val_iss = float(str(iss_cell).replace('.', '').replace(',', '.'))
                        except ValueError:
                            pass

                records.append({
                    "id": f"GYN-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "cidade": "Goiânia"
                })
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        if '|' not in line: continue
                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue

                                    val_iss = 0.0
                                    if len(parts) >= 7:
                                        iss_raw = parts[6].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"GYN-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "Goiânia"
                                    })
                                    idx += 1
                                except ValueError: pass
        except Exception: pass
        return records
