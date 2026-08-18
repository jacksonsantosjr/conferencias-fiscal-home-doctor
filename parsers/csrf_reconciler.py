import pandas as pd
import pdfplumber
import re
from decimal import Decimal, InvalidOperation
from typing import Dict, Any, List

def clean_filial(val):
    if pd.isna(val):
        return ''
    s = str(val).split('.')[0].strip()
    s_digits = re.sub(r'\D', '', s)
    if not s_digits:
        return ''
    s_clean = s_digits.lstrip('0')
    if not s_clean:
        return ''
    if len(s_clean) >= 5:
        empresa = s_clean[:4]
        unidade = s_clean[4:].lstrip('0')
        return f"{empresa}_{unidade}" if unidade else empresa
    return s_clean

def clean_doc_num(val):
    if pd.isna(val):
        return ''
    s = str(val).split('.')[0].strip()
    s_digits = re.sub(r'\D', '', s)
    return s_digits.lstrip('0') if s_digits else s.lstrip('0')

def parse_currency(val_str):
    if not isinstance(val_str, str):
        try:
            return Decimal(str(val_str))
        except:
            return Decimal('0.00')
    val_str = val_str.replace('R$', '').strip()
    if not val_str:
        return Decimal('0.00')
    val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return Decimal(val_str)
    except:
        return Decimal('0.00')

import openpyxl

class CSRFReconciler:
    def __init__(self):
        self.se2_data = []      # PCC from ERP
        self.aglu_data = []     # Aglutinacao
        self.r4020_data = []    # REINF

    def parse_se2(self, file_bytes):
        try:
            wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            coop_cnpjs = set()
            coop_cods = set()
            if 'Cooperativas' in sheet_names:
                ws_coop = wb['Cooperativas']
                header = None
                for row in ws_coop.iter_rows(values_only=True):
                    if not row or all(v is None for v in row): continue
                    if header is None:
                        header = [str(v).strip() if v is not None else '' for v in row]
                        cnpj_idx = header.index('CNPJ') if 'CNPJ' in header else -1
                        forn_idx = header.index('Fornecedor') if 'Fornecedor' in header else -1
                        continue
                    if cnpj_idx != -1 and len(row) > cnpj_idx and row[cnpj_idx]:
                        coop_cnpjs.add(re.sub(r'\D', '', str(row[cnpj_idx]).strip()))
                    if forn_idx != -1 and len(row) > forn_idx and row[forn_idx]:
                        coop_cods.add(str(row[forn_idx]).strip().lstrip('0'))

            forn_dict = {}
            if 'FORNECEDOR' in sheet_names:
                ws_forn = wb['FORNECEDOR']
                for row in ws_forn.iter_rows(values_only=True):
                    if row and len(row) > 1:
                        nome_forn = str(row[0]).strip() if row[0] is not None else ''
                        cnpj_forn = str(row[1]).strip() if row[1] is not None else ''
                        if nome_forn and cnpj_forn:
                            forn_dict[nome_forn.upper()] = cnpj_forn

            if 'SE2' in sheet_names:
                ws_se2 = wb['SE2']
                header_map = None
                for row in ws_se2.iter_rows(values_only=True):
                    if not row or all(v is None for v in row): continue
                    if header_map is None:
                        row_str = ' '.join(str(v).lower() for v in row if v is not None)
                        if 'filial' in row_str and ('titulo' in row_str or 'título' in row_str or 'natureza' in row_str):
                            header_map = {}
                            for idx, val in enumerate(row):
                                if val is not None:
                                    k = str(val).strip()
                                    header_map[k] = idx
                            continue
                        else:
                            continue

                    def get_val(col_name):
                        idx = header_map.get(col_name, -1)
                        return row[idx] if idx != -1 and len(row) > idx and row[idx] is not None else None

                    filial = clean_filial(get_val('Filial'))
                    if not filial: continue

                    numero = clean_doc_num(get_val('No. Titulo'))
                    forn_cod = str(get_val('Fornecedor') or '').strip().lstrip('0')
                    razao = str(get_val('Nome Fornece') or '').strip()
                    cnpj = str(get_val('CNPJ Fornec') or '').strip()

                    if (not cnpj or cnpj == 'nan' or cnpj == '.   .   /    -') and razao.upper() in forn_dict:
                        cnpj = forn_dict[razao.upper()]

                    cnpj_clean = re.sub(r'\D', '', cnpj)
                    pis = parse_currency(get_val('PIS/PASEP') or 0)
                    cofins = parse_currency(get_val('COFINS') or 0)
                    csll = parse_currency(get_val('CSLL') or 0)

                    if pis <= 0 and cofins <= 0 and csll <= 0:
                        continue

                    is_coop = (cnpj_clean in coop_cnpjs and cnpj_clean) or (forn_cod in coop_cods and forn_cod) or ('COOP' in razao.upper())

                    self.se2_data.append({
                        'filial': filial,
                        'numero': numero,
                        'cnpj': cnpj,
                        'razao': razao,
                        'is_coop': 'SIM' if is_coop else 'NÃO',
                        'pis': pis,
                        'cofins': cofins,
                        'csll': csll,
                        'pcc': pis + cofins + csll if not is_coop else Decimal('0.00')
                    })

            wb.close()
        except Exception as e:
            print("Notice: Error in openpyxl streaming parser, using pandas fallback:", e)
            try:
                file_bytes.seek(0)
                xl = pd.ExcelFile(file_bytes)
                coop_cnpjs = set()
                coop_cods = set()
                if 'Cooperativas' in xl.sheet_names:
                    df_coop = pd.read_excel(xl, sheet_name='Cooperativas')
                    df_coop.columns = [str(c).strip() for c in df_coop.columns]
                    if 'CNPJ' in df_coop.columns:
                        coop_cnpjs = set(df_coop['CNPJ'].dropna().astype(str).str.strip().str.replace(r'\D', '', regex=True))
                    if 'Fornecedor' in df_coop.columns:
                        coop_cods = set(df_coop['Fornecedor'].dropna().astype(str).str.strip().str.lstrip('0'))

                forn_dict = {}
                if 'FORNECEDOR' in xl.sheet_names:
                    df_forn = pd.read_excel(xl, sheet_name='FORNECEDOR')
                    for _, row in df_forn.iterrows():
                        nome_forn = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
                        cnpj_forn = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
                        if nome_forn and cnpj_forn:
                            forn_dict[nome_forn.upper()] = cnpj_forn

                header_idx = 1
                df_temp = pd.read_excel(xl, sheet_name='SE2', header=None, nrows=10)
                for i, row in df_temp.iterrows():
                    row_str = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
                    if 'filial' in row_str and ('titulo' in row_str or 'título' in row_str or 'natureza' in row_str):
                        header_idx = i
                        break

                df_se2 = pd.read_excel(xl, sheet_name='SE2', header=header_idx)
                df_se2.columns = [str(c).strip() for c in df_se2.columns]

                for _, row in df_se2.iterrows():
                    filial = clean_filial(row.get('Filial'))
                    if not filial: continue
                    numero = clean_doc_num(row.get('No. Titulo'))
                    forn_cod = str(row.get('Fornecedor', '')).strip().lstrip('0')
                    razao = str(row.get('Nome Fornece', '')).strip() if pd.notna(row.get('Nome Fornece')) else ''
                    cnpj = str(row.get('CNPJ Fornec', '')).strip() if pd.notna(row.get('CNPJ Fornec')) else ''
                    if (not cnpj or cnpj == 'nan' or cnpj == '.   .   /    -') and razao.upper() in forn_dict:
                        cnpj = forn_dict[razao.upper()]
                    cnpj_clean = re.sub(r'\D', '', cnpj)
                    pis = parse_currency(row.get('PIS/PASEP', 0))
                    cofins = parse_currency(row.get('COFINS', 0))
                    csll = parse_currency(row.get('CSLL', 0))
                    if pis <= 0 and cofins <= 0 and csll <= 0: continue
                    is_coop = (cnpj_clean in coop_cnpjs and cnpj_clean) or (forn_cod in coop_cods and forn_cod) or ('COOP' in razao.upper())
                    self.se2_data.append({
                        'filial': filial,
                        'numero': numero,
                        'cnpj': cnpj,
                        'razao': razao,
                        'is_coop': 'SIM' if is_coop else 'NÃO',
                        'pis': pis,
                        'cofins': cofins,
                        'csll': csll,
                        'pcc': pis + cofins + csll if not is_coop else Decimal('0.00')
                    })
            except Exception as e2:
                print("Notice: Could not parse SE2 in fallback either:", e2)

    def parse_aglutinacao(self, file_bytes, is_excel=False):
        try:
            if is_excel:
                try:
                    df = pd.read_excel(file_bytes)
                except Exception as e:
                    print("EXCEL EXCEPTION AGLU:", e)
                    file_bytes.seek(0)
                    df = pd.read_csv(file_bytes)
                
                col_filial = None
                col_numero = None
                col_natureza = None
                col_valor = None
                
                for col in df.columns:
                    col_clean = str(col).lower().replace('ú', 'u').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ã', 'a').replace('ç', 'c').replace('á', 'a').strip()
                    if 'filial' in col_clean: col_filial = col
                    elif 'numero' in col_clean or 'titulo' in col_clean or 'documento' in col_clean or 'tit' in col_clean: col_numero = col
                    elif 'natureza' in col_clean: col_natureza = col
                    elif 'valor' in col_clean or 'vlr' in col_clean: col_valor = col
                
                if not col_natureza:
                    for col in df.columns:
                        col_clean = str(col).lower().replace('ú', 'u').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ã', 'a').replace('ç', 'c').replace('á', 'a').strip()
                        if 'tipo' in col_clean:
                            col_natureza = col
                            break
                
                if not col_filial or not col_numero or not col_valor:
                    print("Aviso: Colunas vitais ausentes no arquivo de Aglutinação.")
                    return
                
                df[col_filial] = df[col_filial].ffill()
                
                for _, row in df.iterrows():
                    filial_raw = str(row[col_filial]) if pd.notna(row[col_filial]) else ''
                    numero_raw = str(row[col_numero]) if pd.notna(row[col_numero]) else ''
                    
                    if not filial_raw or not numero_raw: continue
                    if 'tota' in filial_raw.lower() or not any(c.isdigit() for c in filial_raw) or not any(c.isdigit() for c in numero_raw):
                        continue
                    
                    filial = clean_filial(filial_raw)
                    numero = clean_doc_num(numero_raw)
                    if not filial or not numero: continue
                    
                    natureza_raw = str(row[col_natureza]).upper() if col_natureza and pd.notna(row[col_natureza]) else ''
                    natureza = 'PCC'
                    if 'PIS' in natureza_raw: natureza = 'PIS'
                    elif 'COF' in natureza_raw: natureza = 'COFINS'
                    elif 'CSL' in natureza_raw: natureza = 'CSLL'
                    
                    valor = parse_currency(row[col_valor])
                    
                    self.aglu_data.append({
                        'numero': numero,
                        'filial': filial,
                        'natureza': natureza,
                        'valor': valor
                    })
                return

            with pdfplumber.open(file_bytes) as pdf:
                pattern = r'^(\d+)\s+(\d+)\s+(\d+)\s+([A-Za-z]+)\s+([\d/]+)\s+([\d/]+)\s+([\d/]+)\s+([A-Za-z0-9]+)\s+([\d.,]+)$'
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text: continue
                    print(f"DEBUG AGLU PAGE TEXT HEAD: {text[:200]}")
                    
                    for line in text.split('\n'):
                        line_str = line.strip()
                        if not line_str: continue
                        
                        match = re.match(pattern, line_str)
                        if match:
                            filial = clean_filial(match.group(1))
                            numero = clean_doc_num(match.group(2))
                            natureza_raw = match.group(8).upper()
                            
                            natureza = 'PCC'
                            if 'PIS' in natureza_raw: natureza = 'PIS'
                            elif 'COF' in natureza_raw: natureza = 'COFINS'
                            elif 'CSL' in natureza_raw: natureza = 'CSLL'
                            
                            valor = parse_currency(match.group(9))
                            
                            self.aglu_data.append({
                                'numero': numero,
                                'filial': filial,
                                'natureza': natureza,
                                'valor': valor
                            })
                        else:
                            parts = line_str.split()
                            # Extração permissiva para relatórios em formato de tabela
                            if len(parts) >= 3:
                                digits_parts = [p for p in parts if p.isdigit()]
                                # Exige ao menos filial e documento
                                if len(digits_parts) >= 2:
                                    filial = clean_filial(digits_parts[0])
                                    numero = clean_doc_num(digits_parts[2]) if len(digits_parts) >= 4 else clean_doc_num(digits_parts[1])
                                    
                                    # Procura o último valor monetário na linha
                                    valor = Decimal('0.00')
                                    for p in reversed(parts):
                                        v = parse_currency(p)
                                        if v > 0:
                                            valor = v
                                            break
                                    
                                    # Determina a natureza se possível
                                    natureza = 'PCC'
                                    for kw in ['PIS', 'COF', 'CSL', 'PCC', 'CSRF']:
                                        if kw in line_str.upper():
                                            if kw == 'COF': natureza = 'COFINS'
                                            elif kw == 'CSL': natureza = 'CSLL'
                                            else: natureza = kw
                                            break
                                            
                                    if valor > 0:
                                        self.aglu_data.append({
                                            'numero': numero,
                                            'filial': filial,
                                            'natureza': natureza,
                                            'valor': valor
                                        })
        except Exception as e:
            print("Notice: Could not parse Aglutinacao PDF CSRF:", e)

    def parse_r4020(self, file_bytes):
        try:
            # Lemos os primeiros 15 linhas para tentar achar o cabeçalho verdadeiro
            df_temp = pd.read_excel(file_bytes, header=None, nrows=15)
            header_idx = 0
            for i, row in df_temp.iterrows():
                row_str = ' '.join(str(v).lower() for v in row.values)
                if 'filial' in row_str or 'estabelecimento' in row_str or 'documento' in row_str:
                    header_idx = i
                    break
            
            file_bytes.seek(0)
            df = pd.read_excel(file_bytes, header=header_idx)
            df.columns = [str(c).strip() for c in df.columns]
            
            def find_col(patterns):
                for c in df.columns:
                    c_lower = str(c).lower()
                    if any(p.lower() in c_lower for p in patterns):
                        return c
                return None

            print(f"DEBUG R4020 COLUMNS: {list(df.columns)}")

            filial_col = find_col(['Filial', 'Estabelecimento'])
            cnpj_col = find_col(['CNPJ Participante', 'CNPJ Prestador', 'CNPJ Beneficiário', 'CNPJ Beneficiario'])
            if not cnpj_col:
                # Fallback que exclui coluna da filial
                for c in df.columns:
                    c_low = str(c).lower()
                    if 'cnpj' in c_low and 'filial' not in c_low:
                        cnpj_col = c
                        break
            if not cnpj_col:
                cnpj_col = find_col(['CNPJ'])

            num_col = find_col(['Nº do Documento', 'Documento', 'Num', 'Nº', 'Nota'])
            val_col = find_col(['Valor Agregado'])
            pis_col = find_col(['Valor do PIS', 'Valor PIS'])
            cofins_col = find_col(['Valor do COFINS', 'Valor COFINS'])
            csll_col = find_col(['Valor do CSLL', 'Valor CSLL'])
            nat_col = find_col(['Natureza de Rendimento', 'Natureza'])
            razao_col = find_col(['Nome Beneficiário', 'Nome Beneficiario', 'Beneficiário', 'Beneficiario', 'Razão', 'Razao', 'Nome'])
            tipo_col = find_col(['Tipo'])
            
            # Preenche valores vazios com o valor da linha superior (comum em relatórios do Protheus)
            if filial_col: df[filial_col] = df[filial_col].ffill()
            if cnpj_col: df[cnpj_col] = df[cnpj_col].ffill()
            if num_col: df[num_col] = df[num_col].ffill()
            if razao_col: df[razao_col] = df[razao_col].ffill()
            
            for _, row in df.iterrows():
                if tipo_col:
                    tipo_val = str(row.get(tipo_col, '')).strip().upper()
                    if tipo_val != 'PGT':
                        continue
                        
                val_agregado = parse_currency(row.get(val_col, 0)) if val_col else Decimal('0.00')
                val_pis = parse_currency(row.get(pis_col, 0)) if pis_col else Decimal('0.00')
                val_cofins = parse_currency(row.get(cofins_col, 0)) if cofins_col else Decimal('0.00')
                val_csll = parse_currency(row.get(csll_col, 0)) if csll_col else Decimal('0.00')
                
                nat = str(row.get(nat_col, '')).strip() if nat_col else ''
                razao = str(row.get(razao_col, '')).strip() if razao_col else ''
                cnpj = str(row.get(cnpj_col, '')).strip() if cnpj_col else ''
                filial = clean_filial(row.get(filial_col)) if filial_col else ''
                numero = clean_doc_num(row.get(num_col)) if num_col else ''
                
                if val_agregado > 0:
                    total_val = val_agregado
                else:
                    total_val = val_pis + val_cofins + val_csll

                if total_val <= 0:
                    continue
                
                is_coop = (nat == '15001' or 'COOP' in razao.upper() or (val_pis > 0 and val_csll == 0 and val_agregado == 0))
                
                self.r4020_data.append({
                    'filial': filial,
                    'numero': numero,
                    'cnpj': cnpj,
                    'razao': razao,
                    'is_coop': 'SIM' if is_coop else 'NÃO',
                    'r4020_pis': val_pis,
                    'r4020_cofins': val_cofins,
                    'r4020_csll': val_csll,
                    'valor_agregado': total_val
                })
        except Exception as e:
            print("Notice: Could not parse R-4020 CSRF:", e)

    def reconcile(self):
        has_se2 = len(self.se2_data) > 0
        has_aglu = len(self.aglu_data) > 0
        has_r4020 = len(self.r4020_data) > 0
        num_active = sum([has_se2, has_aglu, has_r4020])

        if num_active < 2:
            return {
                'detalhes': [],
                'resumo': {
                    'total_processado': 0,
                    'conciliados': 0,
                    'divergentes': 0,
                    'ausentes': 0
                },
                'num_relatorios': num_active
            }

        todas_filiais_auxiliares = set()
        for item in self.aglu_data + self.r4020_data:
            f = item.get('filial', '')
            if f:
                todas_filiais_auxiliares.add(f)
        
        if not todas_filiais_auxiliares:
            for item in self.se2_data:
                f = item.get('filial', '')
                if f:
                    todas_filiais_auxiliares.add(f)

        valid_prefixes = set(f.split('_')[0] for f in todas_filiais_auxiliares if f)

        master = {}
        def get_or_create(filial, num):
            key = f"{filial}_{num}"
            if key not in master:
                master[key] = {
                    'filial': filial,
                    'numero': num,
                    'cnpj': '',
                    'razao': '',
                    'is_coop': 'NÃO',
                    'se2_pis': Decimal('0.00'),
                    'se2_cofins': Decimal('0.00'),
                    'se2_pcc': Decimal('0.00'),
                    'aglu_pis': Decimal('0.00'),
                    'aglu_cofins': Decimal('0.00'),
                    'aglu_csll': Decimal('0.00'),
                    'r4020_agregado': Decimal('0.00'),
                    'r4020_pis': Decimal('0.00'),
                    'r4020_cofins': Decimal('0.00'),
                    'r4020_csll': Decimal('0.00'),
                    'status': '',
                    'diagnostico': ''
                }
            return master[key]

        if has_se2:
            for item in self.se2_data:
                f_norm = str(item.get('filial', ''))
                f_prefix = f_norm.split('_')[0] if f_norm else ''
                if valid_prefixes and f_prefix not in valid_prefixes:
                    continue
                
                rec = get_or_create(item['filial'], item['numero'])
                if item.get('cnpj'): rec['cnpj'] = item['cnpj']
                if item.get('razao'): rec['razao'] = item['razao']
                
                # Se já for cooperativa num dos itens da mesma NF, mantém
                if item['is_coop'] == 'SIM': 
                    rec['is_coop'] = 'SIM'
                
                rec['se2_pis'] += item['pis']
                rec['se2_cofins'] += item['cofins']
                rec['se2_pcc'] += item['pcc']

        if has_aglu:
            for item in self.aglu_data:
                rec = get_or_create(item['filial'], item['numero'])
                nat = item['natureza']
                if nat == 'PIS':
                    rec['aglu_pis'] += item['valor']
                elif nat == 'COFINS':
                    rec['aglu_cofins'] += item['valor']
                elif nat in ['CSLL', 'PCC']:
                    rec['aglu_csll'] += item['valor']

        if has_r4020:
            for item in self.r4020_data:
                rec = get_or_create(item['filial'], item['numero'])
                rec['r4020_agregado'] += item['valor_agregado']
                rec['r4020_pis'] += item.get('r4020_pis', Decimal('0.00'))
                rec['r4020_cofins'] += item.get('r4020_cofins', Decimal('0.00'))
                rec['r4020_csll'] += item.get('r4020_csll', Decimal('0.00'))
                if item.get('is_coop') == 'SIM':
                    rec['is_coop'] = 'SIM'
                if item.get('cnpj') and not rec['cnpj']: rec['cnpj'] = item['cnpj']
                if item.get('razao') and not rec['razao']: rec['razao'] = item['razao']

        # Validação de filiais pelo prefixo da cidade (valid_prefixes)
        results = []
        conciliados = 0
        divergentes = 0
        ausentes = 0

        for key, rec in master.items():
            rec_prefix = rec['filial'].split('_')[0] if rec['filial'] else ''
            if valid_prefixes and rec_prefix not in valid_prefixes:
                continue
            
            if rec['is_coop'] == 'SIM':
                se2_total_coop = (rec['se2_pis'] + rec['se2_cofins']) if has_se2 else None
                aglu_total_coop = (rec['aglu_pis'] + rec['aglu_cofins'] + rec['aglu_csll']) if has_aglu else None
                r4020_total_coop = (rec['r4020_pis'] + rec['r4020_cofins'] if (rec['r4020_pis'] > 0 or rec['r4020_cofins'] > 0) else rec['r4020_agregado']) if has_r4020 else None
                
                coop_vals = []
                if has_se2: coop_vals.append(('SE2', se2_total_coop))
                if has_aglu: coop_vals.append(('Aglu.', aglu_total_coop))
                if has_r4020: coop_vals.append(('R-4020', r4020_total_coop))

                non_zero_coop = [v for _, v in coop_vals if v is not None and v > 0]
                zero_coop = [name for name, v in coop_vals if v is not None and v == 0]

                if len(non_zero_coop) == len(coop_vals) and len(non_zero_coop) > 0:
                    if len(set(non_zero_coop)) == 1:
                        rec['status'] = 'Conciliado'
                        if num_active == 3:
                            rec['diagnostico'] = 'Sem divergências (SE2 = Aglu. = R-4020)'
                        elif has_se2 and has_aglu:
                            rec['diagnostico'] = 'Coop. Conciliada (Base SE2 = Aglutinação)'
                        elif has_se2 and has_r4020:
                            rec['diagnostico'] = 'Coop. Conciliada (Base SE2 = REINF R-4020)'
                        elif has_aglu and has_r4020:
                            rec['diagnostico'] = 'Coop. Conciliada (Aglutinação = REINF R-4020)'
                        conciliados += 1
                        diff = Decimal('0.00')
                    else:
                        rec['status'] = 'Divergente'
                        rec['diagnostico'] = 'Divergência de Valores (Coop.)'
                        divergentes += 1
                        diff = max(non_zero_coop) - min(non_zero_coop)
                elif len(non_zero_coop) == 0:
                    continue
                else:
                    rec['status'] = 'Ausente'
                    if has_r4020 and ('R-4020' in zero_coop) and (('SE2' not in zero_coop) or ('Aglu.' not in zero_coop)):
                        rec['diagnostico'] = 'Coop. Ausente no R-4020 (Pendente de Inclusão Manual no REINF)'
                    else:
                        rec['diagnostico'] = f"Coop. Ausente em: {', '.join(zero_coop)}"
                    ausentes += 1
                    diff = max(non_zero_coop) if non_zero_coop else Decimal('0.00')

                rec['valor_erp'] = float(se2_total_coop) if has_se2 else None
                rec['valor_aglu'] = float(aglu_total_coop) if has_aglu else None
                rec['valor_reinf'] = float(r4020_total_coop) if has_r4020 else None
                rec['diferenca'] = float(diff)
                rec['razao'] = f"[COOP] {rec['razao']}" if rec['razao'] and not rec['razao'].startswith('[COOP]') else (rec['razao'] or '[COOP] COOPERATIVA')

            else:
                se2_pcc = rec['se2_pcc'] if has_se2 else None
                aglu_total = (rec['aglu_pis'] + rec['aglu_cofins'] + rec['aglu_csll']) if has_aglu else None
                r4020_total = rec['r4020_agregado'] if has_r4020 else None

                active_vals = []
                if has_se2: active_vals.append(('SE2', se2_pcc))
                if has_aglu: active_vals.append(('Aglu.', aglu_total))
                if has_r4020: active_vals.append(('R-4020', r4020_total))

                non_zero_vals = [v for _, v in active_vals if v is not None and v > 0]
                zero_sources = [name for name, v in active_vals if v is not None and v == 0]

                if len(non_zero_vals) == len(active_vals) and len(non_zero_vals) > 0:
                    if len(set(non_zero_vals)) == 1:
                        rec['status'] = 'Conciliado'
                        if num_active == 3:
                            rec['diagnostico'] = 'Sem divergências (SE2 = Aglu. = R-4020)'
                        elif has_se2 and has_r4020:
                            rec['diagnostico'] = 'Conciliado (Base SE2 = REINF R-4020)'
                        elif has_se2 and has_aglu:
                            rec['diagnostico'] = 'Conciliado (Base SE2 = Aglutinação)'
                        elif has_aglu and has_r4020:
                            rec['diagnostico'] = 'Conciliado (Aglutinação = REINF R-4020)'
                        conciliados += 1
                        diff = Decimal('0.00')
                    else:
                        rec['status'] = 'Divergente'
                        rec['diagnostico'] = 'Divergência de Valores'
                        divergentes += 1
                        diff = max(non_zero_vals) - min(non_zero_vals)
                elif len(non_zero_vals) == 0:
                    continue
                else:
                    rec['status'] = 'Ausente'
                    if has_r4020 and ('R-4020' in zero_sources) and (('SE2' not in zero_sources) or ('Aglu.' not in zero_sources)):
                        rec['diagnostico'] = 'Ausente no R-4020 (Pendente de Inclusão Manual no REINF)'
                    else:
                        rec['diagnostico'] = f"Ausente em: {', '.join(zero_sources)}"
                    ausentes += 1
                    diff = max(non_zero_vals) if non_zero_vals else Decimal('0.00')
                
                rec['valor_erp'] = float(se2_pcc) if has_se2 else None
                rec['valor_aglu'] = float(aglu_total) if has_aglu else None
                rec['valor_reinf'] = float(r4020_total) if has_r4020 else None
                rec['diferenca'] = float(diff)

            rec['valor_erp_fmt'] = f"R$ {rec['valor_erp']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if rec['valor_erp'] is not None else '-'
            rec['valor_aglu_fmt'] = f"R$ {rec['valor_aglu']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if rec['valor_aglu'] is not None else '-'
            rec['valor_reinf_fmt'] = f"R$ {rec['valor_reinf']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if rec['valor_reinf'] is not None else '-'
            rec['diferenca_fmt'] = f"R$ {rec['diferenca']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            results.append(rec)
            
        return {
            'detalhes': results,
            'resumo': {
                'total_processado': len(results),
                'conciliados': conciliados,
                'divergentes': divergentes,
                'ausentes': ausentes
            },
            'num_relatorios': num_active
        }
