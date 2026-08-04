"""
Parser do relatório do ERP (Formato CSV, PDF ou XLSX).
Extrai a coluna BASE DE CÁLCULO / VALOR DOS SERVIÇOS, número do RPS/NFS-e, dados do tomador e datas,
com suporte a planilhas Excel (.xlsx), relatórios PDF, consolidação por número de nota e leitura resiliente com safe_read_bytes.
"""

import csv
import io
import openpyxl
import pdfplumber
from typing import List, Dict, Any
from parsers.base_parser import safe_read_bytes

class ERPParser:
    def parse_file(self, file_source) -> List[Dict[str, Any]]:
        raw_records = self._extract_raw_records(file_source)
        if not raw_records:
            return []

        records = list(raw_records)

        if len(records) > 1:
            total_demais = sum(r["valor"] for r in records[:-1])
            if abs(records[-1]["valor"] - total_demais) < 0.05:
                records.pop()

        return records

    def _extract_raw_records(self, file_source) -> List[Dict[str, Any]]:
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)

        try:
            content = data_bytes.decode('utf-8')
        except Exception:
            content = data_bytes.decode('latin1', errors='replace')

        records = []
        lines = content.splitlines()
        delimiter = ';' if any(';' in line for line in lines[:5]) else ','
        reader = csv.reader(lines, delimiter=delimiter)
        
        headers = None
        idx_situacao = None
        idx_escrituracao = None
        idx_valor = None
        idx_valor_iss = None
        idx_nf = None
        idx_rps = None
        idx_tomador = None
        idx_cnpj = None
        idx_data = None
        
        row_count = 0
        
        for row in reader:
            if not row or len(row) < 3:
                continue
                
            if headers is None:
                headers = row
                for idx, h in enumerate(headers):
                    h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').replace('Ã', 'a').replace('§', 'c').strip()
                    if 'escrituracao' in h_norm:
                        idx_escrituracao = idx
                    elif 'situa' in h_norm:
                        if idx_situacao is None: idx_situacao = idx
                    elif 'imposto devido' in h_norm or ('imposto' in h_norm and 'retido' not in h_norm and idx_valor_iss is None):
                        idx_valor_iss = idx
                    elif 'valor servi' in h_norm or 'valor dos servi' in h_norm or 'valor do servico' in h_norm or ('base' in h_norm and idx_valor is None) or ('valor' in h_norm and idx_valor is None):
                        idx_valor = idx
                    elif 'nr. nf' in h_norm or 'nfs' in h_norm or 'nota' in h_norm:
                        if idx_nf is None: idx_nf = idx
                    elif 'rps' in h_norm:
                        if idx_rps is None: idx_rps = idx
                    elif 'raz' in h_norm or 'tomador' in h_norm or 'nome' in h_norm:
                        if idx_tomador is None and idx > 5: idx_tomador = idx
                    elif 'cnpj' in h_norm or 'cpf' in h_norm:
                        if idx_cnpj is None: idx_cnpj = idx
                    elif 'data' in h_norm and idx_data is None:
                        idx_data = idx
                continue

            row_count += 1

            if idx_escrituracao is not None and idx_escrituracao < len(row):
                esc = row[idx_escrituracao].strip().upper()
                if esc and esc not in ['ATIVA', 'ATIVO', 'T']:
                    continue

            if idx_situacao is not None and idx_situacao < len(row):
                situacao = row[idx_situacao].strip().upper()
                if situacao and situacao not in ['T', 'ATIVA', 'ATIVO', 'TRIBUTADA', 'RETIDA']:
                    continue
            
            val = None
            raw_val = ""

            if idx_valor is not None and idx_valor < len(row):
                raw_val = row[idx_valor].strip()
                if raw_val:
                    try:
                        val = float(raw_val.replace('.', '').replace(',', '.'))
                    except ValueError:
                        pass

            if val is None:
                for col_target in [51, 26, 23, 9]:
                    if len(row) > col_target:
                        raw_val = row[col_target].strip()
                        try:
                            val = float(raw_val.replace('.', '').replace(',', '.'))
                            if val > 0: break
                        except ValueError:
                            pass

            if val is None or val <= 0:
                continue

            # Extração do valor de ISS (IMPOSTO DEVIDO)
            val_iss = 0.0
            if idx_valor_iss is not None and idx_valor_iss < len(row):
                raw_iss = row[idx_valor_iss].strip()
                if raw_iss:
                    try:
                        val_iss = float(raw_iss.replace('.', '').replace(',', '.'))
                    except ValueError:
                        pass

            nf_num = row[idx_nf].strip() if (idx_nf is not None and idx_nf < len(row)) else (row[1].strip() if len(row) > 1 else "")
            rps_num = row[idx_rps].strip() if (idx_rps is not None and idx_rps < len(row)) else (row[2].strip() if len(row) > 2 else "")
            tomador = row[idx_tomador].strip() if (idx_tomador is not None and idx_tomador < len(row)) else (row[7].strip() if len(row) > 7 else "")

            if "TOTAL" in tomador.upper() or "SOMA" in tomador.upper():
                continue

            records.append({
                "id": f"ERP-{len(records)+1}",
                "linha_erp": row_count,
                "nf_num": nf_num,
                "nfs_nac": nf_num,
                "rps": rps_num,
                "valor": val,
                "valor_iss": val_iss,
                "raw_valor": raw_val,
                "tomador": tomador,
                "cnpj_tomador": "",
                "data_emissao": ""
            })

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_escrituracao = None
            idx_situacao = None
            idx_valor = None
            idx_valor_iss = None
            idx_nfs = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'escrituracao' in h_norm:
                    idx_escrituracao = idx
                elif 'situacao' in h_norm:
                    if idx_situacao is None: idx_situacao = idx
                elif 'imposto devido' in h_norm or ('imposto' in h_norm and 'retido' not in h_norm and idx_valor_iss is None):
                    idx_valor_iss = idx
                elif 'valor do servico' in h_norm or 'valor dos servi' in h_norm or ('base' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'nfs' in h_norm or 'nota' in h_norm or 'nr. nf' in h_norm:
                    if idx_nfs is None: idx_nfs = idx
                elif 'nome/razao social' in h_norm or 'razao' in h_norm or 'tomador' in h_norm:
                    if idx > 10: idx_tomador = idx

            if idx_escrituracao is None: idx_escrituracao = 41
            if idx_valor is None: idx_valor = 23
            if idx_nfs is None: idx_nfs = 3
            if idx_tomador is None: idx_tomador = 16

            for row_idx in range(2, sheet.max_row+1):
                esc = sheet.cell(row=row_idx, column=idx_escrituracao+1).value
                sit = sheet.cell(row=row_idx, column=idx_situacao+1).value if idx_situacao is not None else None
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                nfs_cell = sheet.cell(row=row_idx, column=idx_nfs+1).value
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value

                if esc and str(esc).strip().upper() not in ['ATIVA', 'ATIVO', 'T']:
                    continue

                if idx_escrituracao is None and sit and str(sit).strip().upper() not in ['ATIVA', 'ATIVO', 'T', 'TRIBUTADA', 'RETIDA']:
                    continue

                if val_cell is None: continue
                try:
                    val = float(val_cell)
                    if val <= 0: continue

                    nfs = str(int(nfs_cell)) if isinstance(nfs_cell, (int, float)) else str(nfs_cell)

                    # Extração do valor de ISS (IMPOSTO DEVIDO)
                    val_iss = 0.0
                    if idx_valor_iss is not None:
                        iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                        if iss_cell is not None:
                            try:
                                val_iss = float(iss_cell)
                            except (ValueError, TypeError):
                                pass

                    records.append({
                        "id": f"ERP-{len(records)+1}",
                        "linha_erp": row_idx,
                        "nf_num": nfs,
                        "nfs_nac": nfs,
                        "rps": "",
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": str(val_cell),
                        "tomador": str(tomador_cell or ''),
                        "cnpj_tomador": "",
                        "data_emissao": ""
                    })
                except ValueError:
                    pass
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        if '|' in line:
                            parts = [p.strip() for p in line.split('|')]
                            if len(parts) >= 5:
                                dia = parts[1]
                                serie = parts[2]
                                numero = parts[3]
                                base_calc = parts[4]

                                if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                    try:
                                        val = float(base_calc.replace('.', '').replace(',', '.'))
                                        if val <= 0: continue

                                        # Extração do IMPOSTO DEVIDO (parts[6]) para valor_iss
                                        val_iss = 0.0
                                        if len(parts) >= 7:
                                            iss_raw = parts[6].strip()
                                            if iss_raw:
                                                try:
                                                    val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                                except ValueError:
                                                    pass
                                        
                                        num_clean = str(numero).strip()
                                        records.append({
                                            "id": f"ERP-PDF-{idx}",
                                            "linha_erp": page_idx + 1,
                                            "nf_num": num_clean,
                                            "nfs_nac": num_clean,
                                            "rps": "",
                                            "valor": val,
                                            "valor_iss": val_iss,
                                            "raw_valor": base_calc,
                                            "tomador": "Relatório ERP (PDF)",
                                            "cnpj_tomador": "",
                                            "data_emissao": ""
                                        })
                                        idx += 1
                                    except ValueError: pass
                        else:
                            parts = line.split()
                            for p in parts:
                                p_clean = p.replace('.', '').replace(',', '.')
                                try:
                                    val = float(p_clean)
                                    if val > 100.0:
                                        records.append({
                                            "id": f"ERP-PDF-{idx}",
                                            "linha_erp": page_idx + 1,
                                            "nf_num": f"PDF-{idx}",
                                            "nfs_nac": "",
                                            "rps": "",
                                            "valor": val,
                                            "valor_iss": 0.0,
                                            "raw_valor": p,
                                            "tomador": "Relatório ERP (PDF)",
                                            "cnpj_tomador": "",
                                            "data_emissao": ""
                                        })
                                        idx += 1
                                        break
                                except ValueError: pass
            return records
        except Exception:
            return []
