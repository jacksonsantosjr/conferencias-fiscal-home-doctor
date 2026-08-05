"""
Adaptador de Parser para a Prefeitura de Aracaju (SE).
Extrai notas fiscais do relatório oficial da Prefeitura de Aracaju (CSV, XLSX ou PDF),
considerando estritamente a coluna 'Serviço(R$)' e filtrando apenas linhas com Situação == 'Normal'.
"""

import pdfplumber
import openpyxl
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser, safe_read_bytes

def parse_val(val_raw) -> float:
    if isinstance(val_raw, (int, float)):
        return float(val_raw)
    val_str = str(val_raw).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0


class AracajuParser(BaseCityParser):
    def __init__(self):
        super().__init__("Aracaju")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)
        else:
            return self._parse_csv_bytes(data_bytes)

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = ""
            for enc in ['latin1', 'utf-8', 'cp1252']:
                try:
                    content = csv_bytes.decode(enc)
                    break
                except Exception:
                    pass

            lines = content.splitlines()
            if not lines: return []

            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers: return []

            idx_numero = None
            idx_rps = None
            idx_valor = None
            idx_valor_iss = None
            idx_situacao = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('º', '').replace('nº', '').strip()
                if 'situacao' in h_norm and idx_situacao is None:
                    idx_situacao = idx
                elif 'servico(r$)' in h_norm or 'servico' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'issqn(r$)' in h_norm or ('iss' in h_norm and idx_valor_iss is None):
                    idx_valor_iss = idx
                elif 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'rps' in h_norm and idx_rps is None:
                    idx_rps = idx
                elif 'razao social' in h_norm or 'tomador' in h_norm:
                    idx_tomador = idx

            if idx_situacao is None: idx_situacao = 0
            if idx_numero is None: idx_numero = 4
            if idx_rps is None: idx_rps = 6
            if idx_valor is None: idx_valor = 7
            if idx_tomador is None: idx_tomador = 2

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= idx_valor:
                    continue

                # Filtro estrito: apenas Situacao == 'Normal'
                if idx_situacao is not None and idx_situacao < len(row):
                    sit = row[idx_situacao].strip().capitalize()
                    if sit != 'Normal':
                        continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                val = parse_val(raw_val)
                if val <= 0: continue

                num_cell = row[idx_numero].strip() if idx_numero < len(row) else f"AJU-{idx_row+1}"
                rps_cell = row[idx_rps].strip() if idx_rps is not None and idx_rps < len(row) else ""
                
                nf = str(int(num_cell)) if num_cell.isdigit() else num_cell
                rps = str(int(rps_cell)) if rps_cell.isdigit() and rps_cell != '0' else ""
                tomador = row[idx_tomador].strip() if idx_tomador is not None and idx_tomador < len(row) else ""

                val_iss = 0.0
                if idx_valor_iss is not None and idx_valor_iss < len(row):
                    val_iss = parse_val(row[idx_valor_iss].strip())

                records.append({
                    "id": f"AJU-{len(records)+1}",
                    "linha": idx_row + 1,
                    "numero": nf,
                    "rps": rps,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": raw_val,
                    "tomador": tomador,
                    "cidade": "Aracaju"
                })
        except Exception:
            pass

        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_situacao = None
            idx_tomador = None
            idx_retido = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif ('servico(r$)' in h_norm or 'valor servico' in h_norm or 'valor' in h_norm) and idx_valor is None:
                    idx_valor = idx
                elif h_norm == 'iss' or 'issqn' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'iss retido' in h_norm:
                    idx_retido = idx
                elif 'situacao' in h_norm:
                    idx_situacao = idx
                elif 'tomador' in h_norm or 'razao social' in h_norm:
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8 # Valor Serviços
            
            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                sit_cell = sheet.cell(row=row_idx, column=idx_situacao+1).value if idx_situacao is not None else None
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value if idx_tomador is not None else ''

                if sit_cell and str(sit_cell).strip().capitalize() != 'Normal':
                    continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                val_iss = 0.0
                if idx_valor_iss is not None:
                    iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                    if iss_cell is not None:
                        val_iss = parse_val(iss_cell)

                iss_ret_flag = "N"
                if idx_retido is not None:
                    ret_cell = sheet.cell(row=row_idx, column=idx_retido+1).value
                    if ret_cell is not None and str(ret_cell).strip().upper() in ["S", "SIM", "YES", "Y", "1"]:
                        iss_ret_flag = "S"

                records.append({
                    "id": f"AJU-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "iss_retido": iss_ret_flag,
                    "cidade": "Aracaju"
                })
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text: continue
                    is_tomados = 'SERVIÇOS TOMADOS' in text.upper()
                    
                    for line in text.split('\n'):
                        if '|' not in line: continue
                        parts = [p.strip() for p in line.split('|')]
                        
                        if is_tomados and len(parts) >= 10:
                            dia = parts[1]
                            numero = parts[4]
                            valor_docto = parts[5]
                            base_calc = parts[7]
                            iss_retido = parts[9]

                            if dia.isdigit() and numero.isdigit() and valor_docto.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(valor_docto.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue
                                    
                                    val_iss = 0.0
                                    if iss_retido:
                                        try:
                                            val_iss = float(iss_retido.replace('.', '').replace(',', '.'))
                                        except ValueError:
                                            pass

                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"AJU-{idx}",
                                        "pagina": page_idx + 1,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": valor_docto,
                                        "cidade": "Aracaju"
                                    })
                                    idx += 1
                                except ValueError: pass
                        elif not is_tomados and len(parts) >= 9:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue
                                    
                                    val_iss = 0.0
                                    if len(parts) >= 7:
                                        iss_raw = parts[6].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"AJU-{idx}",
                                        "pagina": page_idx + 1,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "Aracaju"
                                    })
                                    idx += 1
                                except ValueError: pass
        except Exception: pass
        return records
