"""
Adaptador de Parser para a Prefeitura de Curitiba (PR).
Extrai notas fiscais do relatório oficial da Prefeitura de Curitiba (XLSX, CSV ou PDF),
considerando estritamente a coluna 'Valor Serviços' (conforme premissa do usuário), com sanitização de texto e descarte de notas canceladas.
"""

import pdfplumber
import openpyxl
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser, safe_read_bytes

def parse_val(val_raw) -> float:
    if isinstance(val_raw, (int, float)):
        return float(val_raw)
    val_str = str(val_raw).replace('R$', '').replace(' ', '').replace('\xa0', '').strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0


class CuritibaParser(BaseCityParser):
    def __init__(self):
        super().__init__("Curitiba")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)
        else:
            return self._parse_csv_bytes(data_bytes)

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_cancelamento = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None:
                    idx_numero = idx
                elif 'valor servicos' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'cancelamento' in h_norm:
                    idx_cancelamento = idx
                elif 'tomador - nome' in h_norm or ('tomador' in h_norm and idx_tomador is None):
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8
            if idx_cancelamento is None: idx_cancelamento = 11
            if idx_tomador is None: idx_tomador = 6

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                canc_cell = sheet.cell(row=row_idx, column=idx_cancelamento+1).value
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value

                if canc_cell and str(canc_cell).strip():
                    continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                records.append({
                    "id": f"CUR-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "cidade": "Curitiba"
                })
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        if '|' not in line: continue
                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue
                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"CUR-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": num_clean,
                                        "valor": val,
                                        "raw_valor": base_calc,
                                        "cidade": "Curitiba"
                                    })
                                    idx += 1
                                except ValueError: pass
        except Exception: pass
        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = csv_bytes.decode('utf-8', errors='replace')
            lines = content.splitlines()
            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers: return []

            idx_numero = None
            idx_valor = None
            idx_cancelamento = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'numero' in h_norm and idx_numero is None: idx_numero = idx
                elif 'valor servicos' in h_norm or ('valor' in h_norm and idx_valor is None): idx_valor = idx
                elif 'cancelamento' in h_norm: idx_cancelamento = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= idx_valor: continue
                if idx_cancelamento is not None and idx_cancelamento < len(row) and row[idx_cancelamento].strip():
                    continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                val = parse_val(raw_val)
                if val <= 0: continue

                num_cell = row[idx_numero].strip() if idx_numero < len(row) else f"CUR-{idx_row+1}"
                nf = str(int(num_cell)) if num_cell.isdigit() else num_cell

                records.append({
                    "id": f"CUR-{len(records)+1}",
                    "linha": idx_row + 1,
                    "numero": nf,
                    "valor": val,
                    "raw_valor": raw_val,
                    "cidade": "Curitiba"
                })
        except Exception: pass
        return records
