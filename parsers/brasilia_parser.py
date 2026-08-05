"""
Adaptador de Parser para a Prefeitura de Brasília (Distrito Federal).
Extrai notas fiscais do Relatório de Serviços Prestados de Brasília (CSV, PDF ou XLSX).
"""

import pdfplumber
import io
import csv
import openpyxl
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser

class BrasiliaParser(BaseCityParser):
    def __init__(self):
        super().__init__("Brasília")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        from parsers.base_parser import safe_read_bytes
        data_bytes = safe_read_bytes(file_source)
        if not data_bytes:
            return []

        if data_bytes.startswith(b'%PDF'):
            return self._parse_pdf_bytes(data_bytes)
        elif data_bytes.startswith(b'PK'):
            return self._parse_xlsx_bytes(data_bytes)
        else:
            return self._parse_csv_bytes(data_bytes)

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column+1)]
            idx_numero = None
            idx_valor = None
            idx_valor_iss = None
            idx_iss_retido = None
            idx_cancelamento = None
            idx_tomador = None

            for idx, h in enumerate(headers):
                if not h: continue
                h_norm = str(h).lower().replace('ã', 'a').replace('ç', 'c').replace('é', 'e').replace('ú', 'u').replace('\n', '').strip()
                if 'iss retido' in h_norm or 'retido' in h_norm:
                    if idx_iss_retido is None: idx_iss_retido = idx
                elif 'numero' in h_norm or 'nmero' in h_norm or ('n' in h_norm and 'mero' in h_norm):
                    if idx_numero is None: idx_numero = idx
                elif 'valor servicos' in h_norm or 'servico(r$)' in h_norm or ('valor' in h_norm and idx_valor is None):
                    idx_valor = idx
                elif h_norm == 'iss' or 'issqn' in h_norm:
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif 'cancelamento' in h_norm:
                    idx_cancelamento = idx
                elif 'tomador - nome' in h_norm or ('tomador' in h_norm and idx_tomador is None):
                    idx_tomador = idx

            if idx_numero is None: idx_numero = 0
            if idx_valor is None: idx_valor = 8
            if idx_cancelamento is None: idx_cancelamento = 11
            if idx_tomador is None: idx_tomador = 6

            def parse_val(v):
                if v is None: return 0.0
                if isinstance(v, (int, float)): return float(v)
                s = str(v).replace('R$', '').replace('.', '').replace(',', '.').strip()
                try:
                    return float(s)
                except ValueError:
                    return 0.0

            for row_idx in range(2, sheet.max_row+1):
                num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value if idx_numero is not None else None
                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value if idx_valor is not None else None
                canc_cell = sheet.cell(row=row_idx, column=idx_cancelamento+1).value if idx_cancelamento is not None else None
                tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value if idx_tomador is not None else None

                if canc_cell and str(canc_cell).strip():
                    continue

                if val_cell is None: continue
                val = parse_val(val_cell)
                if val <= 0: continue

                nf = str(int(num_cell)) if isinstance(num_cell, (int, float)) else str(num_cell).strip()

                val_iss = 0.0
                if idx_valor_iss is not None:
                    iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                    if iss_cell is not None:
                        val_iss = parse_val(iss_cell)

                iss_ret_str = "N"
                if idx_iss_retido is not None:
                    ret_cell = sheet.cell(row=row_idx, column=idx_iss_retido+1).value
                    if ret_cell is not None:
                        ret_val = str(ret_cell).strip().upper()
                        if ret_val in ["SIM", "S", "TRUE", "1"]:
                            iss_ret_str = "S"

                records.append({
                    "id": f"DF-{len(records)+1}",
                    "linha": row_idx,
                    "numero": nf,
                    "valor": val,
                    "valor_iss": val_iss,
                    "iss_retido": iss_ret_str,
                    "raw_valor": str(val_cell),
                    "tomador": str(tomador_cell or ''),
                    "cidade": "Brasília"
                })
        except Exception:
            pass

        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    is_tomados = 'SERVIÇOS TOMADOS' in text.upper() or 'SERVIOS TOMADOS' in text.upper()

                    for line in text.split('\n'):
                        if '|' not in line:
                            continue

                        parts = [p.strip() for p in line.split('|')]
                        if is_tomados and len(parts) >= 10:
                            dia = parts[1]
                            numero = parts[4]
                            valor_docto = parts[5]
                            base_calc = parts[7]
                            iss_retido = parts[9]
                            
                            if dia.isdigit() and numero.isdigit() and valor_docto.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(valor_docto.replace('.', '').replace(',', '.'))
                                    if val <= 0: continue
                                    
                                    val_iss = 0.0
                                    if iss_retido:
                                        try:
                                            val_iss = float(iss_retido.replace('.', '').replace(',', '.'))
                                        except ValueError: pass
                                    
                                    num_clean = str(int(numero))
                                    records.append({
                                        "id": f"DF-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "iss_retido": "N",
                                        "raw_valor": valor_docto,
                                        "cidade": "Brasília"
                                    })
                                    idx += 1
                                except ValueError: pass
                                
                        elif not is_tomados and len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0:
                                        continue

                                    val_iss = 0.0
                                    if len(parts) >= 7:
                                        iss_raw = parts[6].strip()
                                        if iss_raw:
                                            try:
                                                val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                            except ValueError:
                                                pass

                                    num_clean = str(int(numero))

                                    records.append({
                                        "id": f"DF-{idx}",
                                        "pagina": page_idx + 1,
                                        "dia": dia,
                                        "serie": serie,
                                        "numero": num_clean,
                                        "valor": val,
                                        "valor_iss": val_iss,
                                        "raw_valor": base_calc,
                                        "cidade": "Brasília"
                                    })
                                    idx += 1
                                except ValueError:
                                    pass
        except Exception:
            pass

        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = csv_bytes.decode('utf-8', errors='replace')
            lines = content.splitlines()

            header_idx = 0
            for idx, l in enumerate(lines[:10]):
                if 'Natureza' in l or 'Valor Documento' in l or 'Nº' in l or 'N°' in l:
                    header_idx = idx
                    break

            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines[header_idx:], delimiter=delimiter)
            headers = next(reader, None)
            if not headers: return []

            idx_natureza = None
            idx_valor = None
            idx_valor_iss = None
            idx_iss_retido = None
            idx_numero = None
            idx_cnpj = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'natureza' in h_norm:
                    idx_natureza = idx
                elif 'valor documento' in h_norm or ('valor' in h_norm and 'imposto' not in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'imposto retido' in h_norm:
                    idx_iss_retido = idx
                elif 'valor imposto' in h_norm or ('imposto' in h_norm and 'retido' not in h_norm):
                    if idx_valor_iss is None: idx_valor_iss = idx
                elif h.strip() in ['Nº', 'N°', 'NÂ°', 'NÂº', 'Numero', 'Número'] or ('n' in h_norm and 'doc' not in h_norm and idx_numero is None):
                    if idx_numero is None: idx_numero = idx
                elif 'cpf' in h_norm or 'cnpj' in h_norm:
                    if idx_cnpj is None: idx_cnpj = idx

            if idx_natureza is None: idx_natureza = 4
            if idx_valor is None: idx_valor = 5
            if idx_numero is None: idx_numero = 2
            if idx_cnpj is None: idx_cnpj = 3

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= idx_valor: continue

                natureza = row[idx_natureza].strip().upper() if idx_natureza < len(row) else ''

                if not any(k in natureza for k in ['EXIGIVEL', 'EXIGÍVEL', 'TRIBUTAVEL', 'TRIBUTÁVEL', 'TRIBUTADA']):
                    continue

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ''
                if not raw_val: continue

                try:
                    val = float(raw_val.replace('.', '').replace(',', '.'))
                    if val <= 0: continue

                    val_iss = 0.0
                    if idx_valor_iss is not None and idx_valor_iss < len(row):
                        raw_iss = row[idx_valor_iss].strip()
                        if raw_iss:
                            try:
                                val_iss = float(raw_iss.replace('.', '').replace(',', '.'))
                            except ValueError: pass

                    iss_ret_str = "N"
                    if idx_iss_retido is not None and idx_iss_retido < len(row):
                        ret_val = str(row[idx_iss_retido]).strip().upper()
                        if ret_val in ["SIM", "S", "TRUE", "1"]:
                            iss_ret_str = "S"

                    nf = row[idx_numero].strip() if idx_numero < len(row) else f"DF-{idx_row+1}"
                    cnpj = row[idx_cnpj].strip() if idx_cnpj < len(row) else ''

                    records.append({
                        "id": f"DF-{len(records)+1}",
                        "linha": idx_row + 1,
                        "numero": nf,
                        "valor": val,
                        "valor_iss": val_iss,
                        "iss_retido": iss_ret_str,
                        "raw_valor": raw_val,
                        "cnpj_tomador": cnpj,
                        "cidade": "Brasília"
                    })
                except ValueError:
                    pass
        except Exception:
            pass

        return records
