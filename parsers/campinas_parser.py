"""
Adaptador de Parser para a Prefeitura de Campinas.
Extrai notas fiscais do Registro de Notas Fiscais de Serviços Prestados de Campinas (PDF ou CSV).
"""

import pdfplumber
import io
import csv
from typing import List, Dict, Any
from parsers.base_parser import BaseCityParser

class CampinasParser(BaseCityParser):
    def __init__(self):
        super().__init__("Campinas")

    def parse(self, file_source) -> List[Dict[str, Any]]:
        records = []
        
        if isinstance(file_source, bytes):
            if file_source.startswith(b'%PDF'):
                return self._parse_pdf_bytes(file_source)
            elif file_source.startswith(b'PK'):
                return self._parse_xlsx_bytes(file_source)
            else:
                return self._parse_csv_bytes(file_source)
        elif isinstance(file_source, str):
            if file_source.lower().endswith('.pdf'):
                with open(file_source, 'rb') as f:
                    return self._parse_pdf_bytes(f.read())
            elif file_source.lower().endswith('.xlsx') or file_source.lower().endswith('.xls'):
                with open(file_source, 'rb') as f:
                    return self._parse_xlsx_bytes(f.read())
            else:
                with open(file_source, 'rb') as f:
                    return self._parse_csv_bytes(f.read())
        return records

    def _parse_pdf_bytes(self, pdf_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        idx = 1
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if not text:
                        continue

                    for line in text.split('\n'):
                        if '|' not in line:
                            continue

                        parts = [p.strip() for p in line.split('|')]
                        if len(parts) >= 5:
                            dia = parts[1]
                            serie = parts[2]
                            numero = parts[3]
                            base_calc = parts[4]

                            if dia.isdigit() and numero.isdigit() and base_calc.replace('.', '').replace(',', '').isdigit():
                                try:
                                    val = float(base_calc.replace('.', '').replace(',', '.'))
                                    if val <= 0:
                                        continue

                                        val_iss = 0.0
                                        if len(parts) >= 7:
                                            iss_raw = parts[6].strip()
                                            if iss_raw:
                                                try:
                                                    val_iss = float(iss_raw.replace('.', '').replace(',', '.'))
                                                except ValueError:
                                                    pass

                                        num_clean = str(int(numero))

                                        records.append({
                                            "id": f"CAM-{idx}",
                                            "pagina": page_idx + 1,
                                            "dia": dia,
                                            "serie": serie,
                                            "numero": num_clean,
                                            "valor": val,
                                            "valor_iss": val_iss,
                                            "raw_valor": base_calc,
                                            "cidade": "Campinas"
                                        })
                                    idx += 1
                                except ValueError:
                                    pass
        except Exception:
            pass

        return records

    def _parse_csv_bytes(self, csv_bytes: bytes) -> List[Dict[str, Any]]:
        records = []
        try:
            content = csv_bytes.decode('utf-8', errors='replace')
            lines = content.splitlines()
            delimiter = ';' if any(';' in line for line in lines[:5]) else ','
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers: return []

            idx_situacao = None
            idx_valor = None
            idx_valor_iss = None
            idx_numero = None
            idx_tomador = None
            idx_retido = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'situa' in h_norm:
                    idx_situacao = idx
                elif 'valor servi' in h_norm or 'base' in h_norm or ('valor' in h_norm and 'iss' not in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'valor iss' in h_norm or ('iss' in h_norm and idx_valor_iss is None):
                    idx_valor_iss = idx
                elif 'nr. nf' in h_norm or 'nf' in h_norm or 'numero' in h_norm:
                    if idx_numero is None: idx_numero = idx
                elif 'nome empresarial' in h_norm or 'razao' in h_norm or 'tomador' in h_norm:
                    idx_tomador = idx
                elif 'retido' in h_norm:
                    idx_retido = idx

            if idx_situacao is None: idx_situacao = 0
            if idx_valor is None: idx_valor = 9
            if idx_numero is None: idx_numero = 1
            if idx_tomador is None: idx_tomador = 7

            for idx_row, row in enumerate(reader):
                if not row or len(row) <= idx_valor: continue

                if idx_situacao < len(row):
                    sit = row[idx_situacao].strip().upper()
                    if sit and sit not in ['ATIVA', 'ATIVO', 'T', 'TRIBUTADA']:
                        continue

                ret = ""
                if idx_retido is not None and idx_retido < len(row):
                    ret = row[idx_retido].strip()

                raw_val = row[idx_valor].strip() if idx_valor < len(row) else ""
                if not raw_val: continue

                try:
                    val = float(raw_val.replace('.', '').replace(',', '.'))
                    if val <= 0: continue
                    
                    val_iss = 0.0
                    if idx_valor_iss is not None and idx_valor_iss < len(row):
                        raw_iss = row[idx_valor_iss].strip()
                        if raw_iss:
                            try:
                                val_iss = float(raw_iss.replace('.', '').replace(',', '.'))
                            except ValueError: pass

                    num = row[idx_numero].strip() if idx_numero < len(row) else f"CAM-{idx_row+1}"
                    tomador = row[idx_tomador].strip() if idx_tomador < len(row) else ""

                    records.append({
                        "id": f"CAM-{len(records)+1}",
                        "linha": idx_row + 1,
                        "numero": num,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": raw_val,
                        "tomador": tomador,
                        "retido": ret,
                        "iss_retido": "N" if not ret or "não retido" in ret.lower() or "nao retido" in ret.lower() else "S",
                        "cidade": "Campinas"
                    })
                except ValueError:
                    pass
        except Exception:
            pass
        return records

    def _parse_xlsx_bytes(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        import openpyxl
        records = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            sheet = wb.active

            header_row_idx = None
            for r in range(1, 10):
                val = sheet.cell(row=r, column=1).value
                if val and isinstance(val, str) and "PRESTADOR" in val.upper():
                    header_row_idx = r
                    break
            
            if not header_row_idx:
                header_row_idx = 6

            headers = [str(sheet.cell(row=header_row_idx, column=j).value or "") for j in range(1, sheet.max_column+1)]
            
            idx_situacao = None
            idx_valor = None
            idx_valor_iss = None
            idx_numero = None
            idx_tomador = None
            idx_retido = None

            for idx, h in enumerate(headers):
                h_norm = h.lower().replace('ã', 'a').replace('ç', 'c').strip()
                if 'situa' in h_norm:
                    idx_situacao = idx
                elif 'valor servi' in h_norm or 'base' in h_norm or ('valor' in h_norm and 'iss' not in h_norm and idx_valor is None):
                    idx_valor = idx
                elif 'valor iss' in h_norm or ('iss' in h_norm and idx_valor_iss is None):
                    idx_valor_iss = idx
                elif 'nr. nf' in h_norm or 'nf' in h_norm or 'numero' in h_norm:
                    if idx_numero is None: idx_numero = idx
                elif 'nome empresarial' in h_norm or 'razao' in h_norm or 'tomador' in h_norm:
                    idx_tomador = idx
                elif 'retido' in h_norm:
                    idx_retido = idx

            if idx_situacao is None: idx_situacao = 7
            if idx_valor is None: idx_valor = 27
            if idx_numero is None: idx_numero = 5
            if idx_tomador is None: idx_tomador = 12

            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                sit_cell = sheet.cell(row=row_idx, column=idx_situacao+1).value
                if sit_cell:
                    sit = str(sit_cell).strip().upper()
                    if sit and sit not in ['ATIVA', 'ATIVO', 'T', 'TRIBUTADA']:
                        continue

                ret_str = ""
                if idx_retido is not None:
                    ret_cell = sheet.cell(row=row_idx, column=idx_retido+1).value
                    if ret_cell:
                        ret_str = str(ret_cell).strip()

                val_cell = sheet.cell(row=row_idx, column=idx_valor+1).value
                if not val_cell: continue

                try:
                    if isinstance(val_cell, (int, float)):
                        val = float(val_cell)
                    else:
                        val = float(str(val_cell).replace('.', '').replace(',', '.'))
                    
                    if val <= 0: continue
                    
                    val_iss = 0.0
                    if idx_valor_iss is not None:
                        iss_cell = sheet.cell(row=row_idx, column=idx_valor_iss+1).value
                        if iss_cell:
                            try:
                                if isinstance(iss_cell, (int, float)):
                                    val_iss = float(iss_cell)
                                else:
                                    val_iss = float(str(iss_cell).replace('.', '').replace(',', '.'))
                            except ValueError: pass

                    num_cell = sheet.cell(row=row_idx, column=idx_numero+1).value
                    num = str(num_cell).strip() if num_cell else f"CAM-{row_idx}"
                    
                    tomador_cell = sheet.cell(row=row_idx, column=idx_tomador+1).value
                    tomador = str(tomador_cell).strip() if tomador_cell else ""

                    records.append({
                        "id": f"CAM-{len(records)+1}",
                        "linha": row_idx,
                        "numero": num,
                        "valor": val,
                        "valor_iss": val_iss,
                        "raw_valor": str(val_cell),
                        "tomador": tomador,
                        "retido": ret_str,
                        "iss_retido": "N" if not ret_str or "não retido" in ret_str.lower() or "nao retido" in ret_str.lower() else "S",
                        "cidade": "Campinas"
                    })
                except ValueError:
                    pass
        except Exception:
            pass
        return records
